"""
Data reader module for the classification pipeline.

Handles reading Excel and CSV files and extracting the target column
for classification.
"""

import csv
import logging
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)


def _read_descriptions_xlsx(file_path: Path, column_name: str) -> list[str]:
    """Read descriptions from an ``.xlsx`` file."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col_idx = headers.index(column_name)
    except ValueError:
        wb.close()
        raise ValueError(
            f"Column '{column_name}' not found. "
            f"Available columns: {headers}"
        )

    descriptions: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = row[col_idx] if col_idx < len(row) else None
        if value is not None:
            clean_val = str(value).strip()
            if clean_val and clean_val.lower() not in ("none", "null", "nan"):
                descriptions.append(clean_val)

    wb.close()
    return descriptions


def _read_descriptions_csv(file_path: Path, column_name: str) -> list[str]:
    """Read descriptions from a ``.csv`` file."""
    with open(file_path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        if reader.fieldnames is None or column_name not in reader.fieldnames:
            available = list(reader.fieldnames) if reader.fieldnames else []
            raise ValueError(
                f"Column '{column_name}' not found. "
                f"Available columns: {available}"
            )
        descriptions: list[str] = []
        for row in reader:
            value = row.get(column_name)
            if value is not None:
                clean_val = value.strip()
                if clean_val and clean_val.lower() not in ("none", "null", "nan"):
                    descriptions.append(clean_val)
    return descriptions


def read_descriptions(file_path: str | Path, column_name: str) -> list[str]:
    """
    Read descriptions from an Excel or CSV file.

    Parameters
    ----------
    file_path : str | Path
        Path to the ``.xlsx`` or ``.csv`` file.
    column_name : str
        Name of the column to extract (must match header exactly).

    Returns
    -------
    list[str]
        List of description strings (empty cells become ``""``).

    Raises
    ------
    FileNotFoundError
        If ``file_path`` does not exist.
    ValueError
        If ``column_name`` is not found in the header row or
        the file extension is unsupported.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    ext = file_path.suffix.lower()
    if ext == ".csv":
        descriptions = _read_descriptions_csv(file_path, column_name)
    elif ext == ".xlsx":
        descriptions = _read_descriptions_xlsx(file_path, column_name)
    else:
        raise ValueError(f"Unsupported file format: '{ext}' (expected .xlsx or .csv)")

    logger.info(
        "Loaded %d descriptions from '%s' (column: '%s')",
        len(descriptions),
        file_path.name,
        column_name,
    )
    return descriptions
