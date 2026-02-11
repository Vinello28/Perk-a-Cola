"""
Streamlit GUI for the LLM classification pipeline.

Launch with::

    streamlit run app/src/gui.py
"""

from __future__ import annotations

import asyncio
import io
import logging
import tempfile
import time
from pathlib import Path
from threading import Thread

import openpyxl
import pandas as pd
import streamlit as st
import yaml

from classifier import LMStudioClassifier
from config import (
    ClassificationConfig,
    ConcurrencyConfig,
    LLMConfig,
    PromptConfig,
    load_config,
)
from output_writer import write_results

# ── Constants ────────────────────────────────────────────────────────

_CONFIG_PATH = Path(__file__).parent / "config.yaml"
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s │ %(levelname)-8s │ %(name)s │ %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("gui")


# ── Helpers ──────────────────────────────────────────────────────────


def _read_headers(file_bytes: bytes) -> list[str]:
    """Read column headers from an in-memory Excel file."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value is not None]
    wb.close()
    return headers


def _read_column(file_bytes: bytes, column_name: str) -> list[str]:
    """Read one column from an in-memory Excel file (skip header)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = headers.index(column_name)
    descriptions: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = row[col_idx] if col_idx < len(row) else None
        descriptions.append(str(value).strip() if value is not None else "")
    wb.close()
    return descriptions


def _load_raw_yaml() -> dict:
    """Load the raw YAML config for display purposes."""
    with open(_CONFIG_PATH, "r", encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def _run_classification_in_thread(
    descriptions: list[str],
    cfg: dict,
    system_prompt_override: str,
    result_container: dict,
) -> None:
    """Run the async classification in a dedicated thread with its own event loop."""

    async def _inner():
        llm_cfg_raw = cfg.get("llm", {})
        cls_cfg_raw = cfg.get("classification", {})
        prompt_cfg_raw = cfg.get("prompt", {})
        concurrency_cfg_raw = cfg.get("concurrency", {})

        llm_cfg = LLMConfig(**{k: v for k, v in llm_cfg_raw.items() if k in LLMConfig.__dataclass_fields__})
        cls_cfg = ClassificationConfig(**{k: v for k, v in cls_cfg_raw.items() if k in ClassificationConfig.__dataclass_fields__})
        concurrency_cfg = ConcurrencyConfig(**{k: v for k, v in concurrency_cfg_raw.items() if k in ConcurrencyConfig.__dataclass_fields__})

        # Override the system prompt with the user-edited version
        prompt_cfg = PromptConfig(system=system_prompt_override, user=prompt_cfg_raw.get("user", ""))

        classifier = LMStudioClassifier(
            llm_cfg=llm_cfg,
            cls_cfg=cls_cfg,
            prompt_cfg=prompt_cfg,
            concurrency_cfg=concurrency_cfg,
        )

        def on_progress(completed: int, total: int):
            result_container["completed"] = completed
            result_container["total"] = total

        labels = await classifier.classify_batch(descriptions, progress_callback=on_progress)
        result_container["labels"] = labels
        result_container["done"] = True

    loop = asyncio.new_event_loop()
    try:
        loop.run_until_complete(_inner())
    except Exception as exc:
        result_container["error"] = str(exc)
        result_container["done"] = True
    finally:
        loop.close()


# ── Page config ──────────────────────────────────────────────────────

st.set_page_config(
    page_title="Perk-a-Cola · LLM Classifier",
    page_icon="🧪",
    layout="wide",
)

# ── Custom CSS ───────────────────────────────────────────────────────

st.markdown("""
<style>
    /* Main container */
    .stApp {
        background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%);
    }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: rgba(15, 12, 41, 0.95);
        border-right: 1px solid rgba(139, 92, 246, 0.3);
    }

    /* Cards */
    .metric-card {
        background: rgba(255, 255, 255, 0.05);
        border: 1px solid rgba(139, 92, 246, 0.25);
        border-radius: 12px;
        padding: 1.2rem;
        text-align: center;
        backdrop-filter: blur(10px);
        transition: transform 0.2s ease, border-color 0.2s ease;
    }
    .metric-card:hover {
        transform: translateY(-2px);
        border-color: rgba(139, 92, 246, 0.6);
    }
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        background: linear-gradient(135deg, #a78bfa, #7c3aed);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        line-height: 1.2;
    }
    .metric-label {
        font-size: 0.85rem;
        color: rgba(255, 255, 255, 0.6);
        margin-top: 0.3rem;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    /* Header */
    .app-header {
        text-align: center;
        padding: 1rem 0 2rem;
    }
    .app-header h1 {
        background: linear-gradient(135deg, #c084fc, #7c3aed, #a78bfa);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 2.5rem;
        font-weight: 800;
        margin-bottom: 0.3rem;
    }
    .app-header p {
        color: rgba(255, 255, 255, 0.5);
        font-size: 1rem;
    }

    /* Sidebar section titles */
    .sidebar-section {
        font-size: 0.75rem;
        font-weight: 600;
        color: rgba(139, 92, 246, 0.8);
        text-transform: uppercase;
        letter-spacing: 0.1em;
        margin: 1.5rem 0 0.5rem;
        padding-bottom: 0.3rem;
        border-bottom: 1px solid rgba(139, 92, 246, 0.2);
    }

    /* Status badge */
    .status-badge {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 9999px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    .status-ready {
        background: rgba(34, 197, 94, 0.15);
        color: #4ade80;
        border: 1px solid rgba(34, 197, 94, 0.3);
    }
    .status-running {
        background: rgba(234, 179, 8, 0.15);
        color: #facc15;
        border: 1px solid rgba(234, 179, 8, 0.3);
    }
    .status-done {
        background: rgba(139, 92, 246, 0.15);
        color: #a78bfa;
        border: 1px solid rgba(139, 92, 246, 0.3);
    }

    /* Progress bar override */
    .stProgress > div > div {
        background: linear-gradient(90deg, #7c3aed, #a78bfa) !important;
        border-radius: 6px;
    }

    /* Hide default streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ── Header ───────────────────────────────────────────────────────────

st.markdown("""
<div class="app-header">
    <h1>🧪 Perk-a-Cola</h1>
    <p>LLM-powered text classification pipeline</p>
</div>
""", unsafe_allow_html=True)

# ── Load config ──────────────────────────────────────────────────────

raw_cfg = _load_raw_yaml()

# ── Sidebar ──────────────────────────────────────────────────────────

with st.sidebar:
    st.image("https://img.shields.io/badge/Perk--a--Cola-LLM_Classifier-7c3aed?style=for-the-badge", use_container_width=True)

    # ── File upload ──────────────────────────────────────────────
    st.markdown('<div class="sidebar-section">📂 Input Data</div>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "Upload an Excel file (.xlsx)",
        type=["xlsx"],
        help="Select the Excel file containing the descriptions to classify.",
    )

    selected_column: str | None = None
    headers: list[str] = []

    if uploaded_file is not None:
        file_bytes = uploaded_file.getvalue()
        headers = _read_headers(file_bytes)
        selected_column = st.selectbox(
            "Select the column to classify",
            options=headers,
            index=headers.index(raw_cfg.get("classification", {}).get("description_column", headers[0]))
            if raw_cfg.get("classification", {}).get("description_column", "") in headers
            else 0,
            help="Choose which column contains the text descriptions.",
        )

    # ── System prompt ────────────────────────────────────────────
    st.markdown('<div class="sidebar-section">💬 System Prompt</div>', unsafe_allow_html=True)
    default_system_prompt = raw_cfg.get("prompt", {}).get("system", "")
    system_prompt = st.text_area(
        "System prompt (editable)",
        value=default_system_prompt,
        height=220,
        help="This prompt is sent to the LLM as the system message. Edit it to customise classification behaviour.",
    )

    # ── LLM info ─────────────────────────────────────────────────
    st.markdown('<div class="sidebar-section">⚙️ Model Configuration</div>', unsafe_allow_html=True)
    llm_info = raw_cfg.get("llm", {})
    st.markdown(f"""
    | Parameter | Value |
    | :--- | :--- |
    | **Model** | `{llm_info.get('model_name', 'N/A')}` |
    | **Base URL** | `{llm_info.get('base_url', 'N/A')}` |
    | **Temperature** | `{llm_info.get('temperature', 'N/A')}` |
    | **Max tokens** | `{llm_info.get('max_tokens', 'N/A')}` |
    | **Thinking** | `{llm_info.get('enable_thinking', 'N/A')}` |
    | **Workers** | `{raw_cfg.get('concurrency', {}).get('max_workers', 'N/A')}` |
    """)

    # ── Labels ───────────────────────────────────────────────────
    st.markdown('<div class="sidebar-section">🏷️ Labels</div>', unsafe_allow_html=True)
    labels = raw_cfg.get("classification", {}).get("labels", [])
    for lbl in labels:
        st.code(lbl, language=None)

    # ── Run button ───────────────────────────────────────────────
    st.markdown("---")
    run_disabled = uploaded_file is None or selected_column is None
    run_button = st.button(
        "🚀 Avvia Classificazione",
        use_container_width=True,
        disabled=run_disabled,
        type="primary",
    )

# ── Main area ────────────────────────────────────────────────────────

if uploaded_file is None:
    # Empty state
    st.markdown("""
    <div style="text-align: center; padding: 4rem 0; color: rgba(255,255,255,0.4);">
        <p style="font-size: 3rem; margin-bottom: 0.5rem;">📂</p>
        <p style="font-size: 1.1rem;">Carica un file Excel dalla sidebar per iniziare</p>
    </div>
    """, unsafe_allow_html=True)

elif not run_button and "results_df" not in st.session_state:
    # File uploaded but not running yet – show preview
    st.subheader("📋 Anteprima del file")
    preview_df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
    st.dataframe(preview_df.head(20), use_container_width=True, height=400)
    st.caption(f"Mostrando le prime 20 righe di **{len(preview_df)}** totali · Colonna selezionata: **{selected_column}**")

elif run_button:
    # ── Classification run ───────────────────────────────────────
    descriptions = _read_column(file_bytes, selected_column)
    total = len(descriptions)

    st.markdown(f"""
    <div style="text-align: center; margin-bottom: 1rem;">
        <span class="status-badge status-running">⏳ Classificazione in corso</span>
    </div>
    """, unsafe_allow_html=True)

    # Metric cards row
    col1, col2, col3 = st.columns(3)
    metric_completed = col1.empty()
    metric_throughput = col2.empty()
    metric_elapsed = col3.empty()

    # Progress bar
    progress_bar = st.progress(0.0)
    progress_text = st.empty()

    # ── Shared state for the background thread ───────────────────
    result_container: dict = {"completed": 0, "total": total, "done": False, "labels": None, "error": None}

    thread = Thread(
        target=_run_classification_in_thread,
        args=(descriptions, raw_cfg, system_prompt, result_container),
        daemon=True,
    )

    t0 = time.perf_counter()
    thread.start()

    # Poll progress
    while not result_container["done"]:
        time.sleep(0.3)
        completed = result_container["completed"]
        elapsed = time.perf_counter() - t0
        throughput = completed / elapsed if elapsed > 0 else 0.0
        fraction = completed / total if total > 0 else 0.0

        progress_bar.progress(fraction)
        progress_text.caption(f"**{completed}** / **{total}** descrizioni classificate")

        metric_completed.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{completed}/{total}</div>
            <div class="metric-label">Completate</div>
        </div>
        """, unsafe_allow_html=True)

        metric_throughput.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{throughput:.1f}</div>
            <div class="metric-label">desc / sec</div>
        </div>
        """, unsafe_allow_html=True)

        metric_elapsed.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{elapsed:.1f}s</div>
            <div class="metric-label">Tempo trascorso</div>
        </div>
        """, unsafe_allow_html=True)

    thread.join()
    elapsed_final = time.perf_counter() - t0
    throughput_final = total / elapsed_final if elapsed_final > 0 else 0.0

    # Final metrics update
    progress_bar.progress(1.0)
    progress_text.caption(f"**{total}** / **{total}** descrizioni classificate")

    metric_completed.markdown(f"""
    <div class="metric-card">
        <div class="metric-value">{total}/{total}</div>
        <div class="metric-label">Completate</div>
    </div>
    """, unsafe_allow_html=True)

    metric_throughput.markdown(f"""
    <div class="metric-card">
        <div class="metric-value">{throughput_final:.1f}</div>
        <div class="metric-label">desc / sec</div>
    </div>
    """, unsafe_allow_html=True)

    metric_elapsed.markdown(f"""
    <div class="metric-card">
        <div class="metric-value">{elapsed_final:.1f}s</div>
        <div class="metric-label">Tempo trascorso</div>
    </div>
    """, unsafe_allow_html=True)

    if result_container.get("error"):
        st.error(f"❌ Errore durante la classificazione: {result_container['error']}")
    else:
        labels = result_container["labels"]
        st.markdown("""
        <div style="text-align: center; margin: 1rem 0;">
            <span class="status-badge status-done">✅ Classificazione completata</span>
        </div>
        """, unsafe_allow_html=True)

        # Build results DataFrame
        results_df = pd.DataFrame({"Descrizione": descriptions, "Label": labels})
        st.session_state["results_df"] = results_df

        # Results table
        st.subheader("📊 Risultati")

        # Label distribution
        dist_col1, dist_col2 = st.columns([1, 2])
        with dist_col1:
            label_counts = results_df["Label"].value_counts()
            st.dataframe(
                label_counts.rename("Conteggio").to_frame(),
                use_container_width=True,
            )
        with dist_col2:
            st.bar_chart(results_df["Label"].value_counts(), color="#7c3aed")

        st.dataframe(results_df, use_container_width=True, height=400)

        # Download button
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)
            write_results(descriptions, labels, tmp_path)
            excel_bytes = tmp_path.read_bytes()

        st.download_button(
            label="⬇️  Scarica risultati (.xlsx)",
            data=excel_bytes,
            file_name=f"{uploaded_file.name.replace('.xlsx', '')}_classificata.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

# Show previous results if available
elif "results_df" in st.session_state:
    results_df = st.session_state["results_df"]
    st.markdown("""
    <div style="text-align: center; margin: 1rem 0;">
        <span class="status-badge status-done">✅ Risultati precedenti</span>
    </div>
    """, unsafe_allow_html=True)

    st.subheader("📊 Risultati")
    dist_col1, dist_col2 = st.columns([1, 2])
    with dist_col1:
        st.dataframe(
            results_df["Label"].value_counts().rename("Conteggio").to_frame(),
            use_container_width=True,
        )
    with dist_col2:
        st.bar_chart(results_df["Label"].value_counts(), color="#7c3aed")

    st.dataframe(results_df, use_container_width=True, height=400)
